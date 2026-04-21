"""
Роутер для генерации отчетов в форматах PDF и XLSX.
Поддерживает индивидуальные отчеты по заявкам, сводные отчеты и отчёт по выполненным заявкам.
"""

import io
import os
from datetime import datetime, timedelta
from typing import Annotated, List, Optional

from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlmodel import select
from sqlalchemy.orm import selectinload

# Генерация PDF
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import inch
from reportlab.platypus import (
    Image,
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)

# Генерация XLSX
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from database import get_session, get_ticket_by_id
from models import DigestRange, Ticket, TicketStatus, User, UserRole
from routers.auth import check_admin_role, get_current_active_user

# ===== КОНФИГУРАЦИЯ =====

router = APIRouter(prefix="/api/reports", tags=["reports"])

# Стили для Excel
EXCEL_HEADER_FONT = Font(bold=True, size=12)
EXCEL_HEADER_ALIGNMENT = Alignment(horizontal="center")
EXCEL_HEADER_FILL = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
EXCEL_LABEL_FONT = Font(bold=True)


# ===== ВСПОМОГАТЕЛЬНЫЕ ФУНКЦИИ ДЛЯ EXCEL =====

def auto_adjust_column_width(worksheet, max_width: int = 50) -> None:
    """Автоматическая настройка ширины столбцов по содержимому."""
    for column in worksheet.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if cell.value and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except Exception:
                pass
        adjusted_width = min(max_length + 2, max_width)
        worksheet.column_dimensions[column_letter].width = adjusted_width


def apply_header_style(worksheet, row_number: int = 1) -> None:
    """Применяет стиль заголовков к строке."""
    for cell in worksheet[row_number]:
        cell.font = EXCEL_HEADER_FONT
        cell.alignment = EXCEL_HEADER_ALIGNMENT
        cell.fill = EXCEL_HEADER_FILL


# ===== ГЕНЕРАЦИЯ PDF =====

def create_ticket_pdf(ticket: Ticket) -> io.BytesIO:
    """Создает PDF отчет для одной заявки."""
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    styles = getSampleStyleSheet()

    title_style = ParagraphStyle(
        "CustomTitle",
        parent=styles["Heading1"],
        fontSize=16,
        spaceAfter=30,
        alignment=1,
    )

    normal_style = ParagraphStyle(
        "CustomNormal",
        parent=styles["Normal"],
        fontSize=12,
        spaceAfter=12,
    )

    story = []
    story.append(Paragraph(f"ОТЧЕТ ПО ЗАЯВКЕ №{ticket.id}", title_style))
    story.append(Spacer(1, 20))

    # Основная информация
    info_data = [
        ["Заголовок:", ticket.title],
        ["Адрес:", ticket.address],
        ["Описание:", ticket.description],
        ["Статус:", ticket.status.value],
        ["Приоритет:", f"{ticket.priority}/5"],
        ["Дата создания:", ticket.created_at.strftime("%d.%m.%Y %H:%M")],
        ["Крайний срок:", ticket.deadline.strftime("%d.%m.%Y %H:%M")],
    ]

    if ticket.customer:
        info_data.append(["Заказчик:", ticket.customer.full_name])
    if ticket.executor:
        info_data.append(["Исполнитель:", ticket.executor.full_name])

    if ticket.started_at:
        info_data.append(["Начата:", ticket.started_at.strftime("%d.%m.%Y %H:%M")])
    if ticket.completed_at:
        info_data.append(["Завершена:", ticket.completed_at.strftime("%d.%m.%Y %H:%M")])
        if ticket.started_at:
            duration = ticket.completed_at - ticket.started_at
            hours = duration.total_seconds() / 3600
            info_data.append(["Время выполнения:", f"{hours:.1f} часов"])

    if ticket.completion_comment:
        info_data.append(["Комментарий:", ticket.completion_comment])
    if ticket.rejection_reason:
        info_data.append(["Причина отклонения:", ticket.rejection_reason])

    table = Table(info_data, colWidths=[2 * inch, 4 * inch])
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (0, -1), colors.lightgrey),
                ("TEXTCOLOR", (0, 0), (-1, -1), colors.black),
                ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                ("FONTNAME", (0, 0), (-1, -1), "Helvetica"),
                ("FONTSIZE", (0, 0), (-1, -1), 10),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 12),
                ("BACKGROUND", (1, 0), (1, -1), colors.beige),
                ("GRID", (0, 0), (-1, -1), 1, colors.black),
            ]
        )
    )
    story.append(table)

    # Информация о фотографиях
    if ticket.before_photo_path or ticket.after_photo_path:
        story.append(Spacer(1, 20))
        story.append(Paragraph("ПРИЛОЖЕННЫЕ ФОТОГРАФИИ:", styles["Heading2"]))

        photos_info = []
        if ticket.before_photo_path:
            photos_info.append(f"• Фото 'до': {os.path.basename(ticket.before_photo_path)}")
        if ticket.after_photo_path:
            photos_info.append(f"• Фото 'после': {os.path.basename(ticket.after_photo_path)}")

        for photo_info in photos_info:
            story.append(Paragraph(photo_info, normal_style))

    # Подвал
    story.append(Spacer(1, 30))
    story.append(
        Paragraph(
            f"Отчет сгенерирован: {datetime.now().strftime('%d.%m.%Y %H:%M')}",
            styles["Normal"],
        )
    )

    doc.build(story)
    buffer.seek(0)
    return buffer


def create_digest_pdf(tickets: List[Ticket], period: str) -> io.BytesIO:
    """Создает сводный PDF отчет за период."""
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    styles = getSampleStyleSheet()
    story = []

    period_text = "дневной" if period == "daily" else "недельный"
    title = f"СВОДНЫЙ {period_text.upper()} ОТЧЕТ"
    story.append(Paragraph(title, styles["Title"]))
    story.append(Spacer(1, 20))

    # Общая статистика
    total_tickets = len(tickets)
    status_counts = {status: 0 for status in TicketStatus}
    for ticket in tickets:
        status_counts[ticket.status] += 1

    stats_data = [
        ["Показатель", "Значение"],
        ["Всего заявок:", str(total_tickets)],
        ["Ожидающих:", str(status_counts[TicketStatus.PENDING])],
        ["В работе:", str(status_counts[TicketStatus.IN_PROGRESS])],
        ["Выполненных:", str(status_counts[TicketStatus.DONE])],
        ["Отклоненных:", str(status_counts[TicketStatus.REJECTED])],
    ]

    completed_tickets = [
        t
        for t in tickets
        if t.status == TicketStatus.DONE and t.started_at and t.completed_at
    ]
    if completed_tickets:
        avg_time = (
            sum((t.completed_at - t.started_at).total_seconds() for t in completed_tickets)
            / len(completed_tickets)
            / 3600
        )
        stats_data.append(["Среднее время выполнения:", f"{avg_time:.1f} часов"])

    stats_table = Table(stats_data, colWidths=[3 * inch, 2 * inch])
    stats_table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.grey),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                ("FONTNAME", (0, 0), (-1, -1), "Helvetica"),
                ("FONTSIZE", (0, 0), (-1, -1), 10),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 12),
                ("BACKGROUND", (0, 1), (-1, -1), colors.beige),
                ("GRID", (0, 0), (-1, -1), 1, colors.black),
            ]
        )
    )
    story.append(Paragraph("ОБЩАЯ СТАТИСТИКА", styles["Heading2"]))
    story.append(stats_table)
    story.append(Spacer(1, 20))

    # Список заявок
    if tickets:
        story.append(Paragraph("СПИСОК ЗАЯВОК", styles["Heading2"]))

        tickets_data = [["№", "Заголовок", "Статус", "Исполнитель", "Дата создания"]]
        for ticket in tickets[:20]:
            executor_name = ticket.executor.full_name if ticket.executor else "Не назначен"
            title_short = (
                ticket.title[:30] + "..." if len(ticket.title) > 30 else ticket.title
            )
            executor_short = (
                executor_name[:20] + "..." if len(executor_name) > 20 else executor_name
            )
            tickets_data.append(
                [
                    str(ticket.id),
                    title_short,
                    ticket.status.value,
                    executor_short,
                    ticket.created_at.strftime("%d.%m.%Y"),
                ]
            )

        tickets_table = Table(
            tickets_data,
            colWidths=[0.5 * inch, 2.5 * inch, 1 * inch, 1.5 * inch, 1 * inch],
        )
        tickets_table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.grey),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                    ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                    ("FONTNAME", (0, 0), (-1, -1), "Helvetica"),
                    ("FONTSIZE", (0, 0), (-1, -1), 8),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
                    ("BACKGROUND", (0, 1), (-1, -1), colors.beige),
                    ("GRID", (0, 0), (-1, -1), 1, colors.black),
                ]
            )
        )
        story.append(tickets_table)

    story.append(Spacer(1, 30))
    story.append(
        Paragraph(
            f"Отчет сгенерирован: {datetime.now().strftime('%d.%m.%Y %H:%M')}",
            styles["Normal"],
        )
    )

    doc.build(story)
    buffer.seek(0)
    return buffer


# ===== ГЕНЕРАЦИЯ XLSX =====

def create_ticket_xlsx(ticket: Ticket) -> io.BytesIO:
    """Создает XLSX отчет для одной заявки."""
    buffer = io.BytesIO()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Заявка №{ticket.id}"

    # Заголовок
    ws.merge_cells("A1:B1")
    ws["A1"] = f"ОТЧЕТ ПО ЗАЯВКЕ №{ticket.id}"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].fill = EXCEL_HEADER_FILL
    ws["A1"].alignment = Alignment(horizontal="center")

    row = 3
    data_pairs = [
        ("Заголовок", ticket.title),
        ("Адрес", ticket.address),
        ("Описание", ticket.description),
        ("Статус", ticket.status.value),
        ("Приоритет", f"{ticket.priority}/5"),
        ("Дата создания", ticket.created_at.strftime("%d.%m.%Y %H:%M")),
        ("Крайний срок", ticket.deadline.strftime("%d.%m.%Y %H:%M")),
    ]

    if ticket.customer:
        data_pairs.append(("Заказчик", ticket.customer.full_name))
    if ticket.executor:
        data_pairs.append(("Исполнитель", ticket.executor.full_name))
    if ticket.started_at:
        data_pairs.append(("Начата", ticket.started_at.strftime("%d.%m.%Y %H:%M")))
    if ticket.completed_at:
        data_pairs.append(("Завершена", ticket.completed_at.strftime("%d.%m.%Y %H:%M")))
        if ticket.started_at:
            duration = ticket.completed_at - ticket.started_at
            hours = duration.total_seconds() / 3600
            data_pairs.append(("Время выполнения", f"{hours:.1f} часов"))

    if ticket.completion_comment:
        data_pairs.append(("Комментарий", ticket.completion_comment))
    if ticket.rejection_reason:
        data_pairs.append(("Причина отклонения", ticket.rejection_reason))

    for label, value in data_pairs:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = EXCEL_LABEL_FONT
        ws[f"B{row}"] = value
        row += 1

    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 50

    wb.save(buffer)
    buffer.seek(0)
    return buffer


def create_digest_xlsx(tickets: List[Ticket], period: str) -> io.BytesIO:
    """Создает сводный XLSX отчет за период."""
    buffer = io.BytesIO()
    wb = openpyxl.Workbook()

    # Лист статистики
    ws_stats = wb.active
    ws_stats.title = "Статистика"

    ws_stats["A1"] = f"СВОДНЫЙ {'ДНЕВНОЙ' if period == 'daily' else 'НЕДЕЛЬНЫЙ'} ОТЧЕТ"
    ws_stats["A1"].font = Font(bold=True, size=14)
    ws_stats.merge_cells("A1:B1")

    total_tickets = len(tickets)
    status_counts = {status: 0 for status in TicketStatus}
    for ticket in tickets:
        status_counts[ticket.status] += 1

    stats_data = [
        ("Всего заявок", total_tickets),
        ("Ожидающих", status_counts[TicketStatus.PENDING]),
        ("В работе", status_counts[TicketStatus.IN_PROGRESS]),
        ("Выполненных", status_counts[TicketStatus.DONE]),
        ("Отклоненных", status_counts[TicketStatus.REJECTED]),
    ]

    row = 3
    for label, value in stats_data:
        ws_stats[f"A{row}"] = label
        ws_stats[f"A{row}"].font = EXCEL_LABEL_FONT
        ws_stats[f"B{row}"] = value
        row += 1

    # Лист с заявками
    ws_tickets = wb.create_sheet("Заявки")
    headers = [
        "ID",
        "Заголовок",
        "Адрес",
        "Статус",
        "Приоритет",
        "Заказчик",
        "Исполнитель",
        "Дата создания",
        "Завершена",
    ]
    ws_tickets.append(headers)
    apply_header_style(ws_tickets)

    for ticket in tickets:
        row_data = [
            ticket.id,
            ticket.title,
            ticket.address,
            ticket.status.value,
            ticket.priority,
            ticket.customer.full_name if ticket.customer else "",
            ticket.executor.full_name if ticket.executor else "",
            ticket.created_at.strftime("%d.%m.%Y %H:%M"),
            ticket.completed_at.strftime("%d.%m.%Y %H:%M") if ticket.completed_at else "",
        ]
        ws_tickets.append(row_data)

    auto_adjust_column_width(ws_tickets)

    wb.save(buffer)
    buffer.seek(0)
    return buffer


def create_executed_xlsx(tickets: List[Ticket]) -> io.BytesIO:
    """Создает Excel отчёт по выполненным заявкам."""
    buffer = io.BytesIO()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Выполненные заявки"

    headers = ["Номер объекта", "Система", "Комментарий", "Исполнитель", "Время выполнения"]
    ws.append(headers)
    apply_header_style(ws)

    for ticket in tickets:
        # Поле 'system' может отсутствовать в модели, используем getattr для безопасности
        system_value = getattr(ticket, "system", "") or ""
        comment = ticket.completion_comment or ticket.description or ""
        executor_name = ticket.executor.full_name if ticket.executor else ""
        completed_str = ticket.completed_at.strftime("%d.%m.%Y %H:%M") if ticket.completed_at else ""
        row = [ticket.title, system_value, comment, executor_name, completed_str]
        ws.append(row)

    auto_adjust_column_width(ws)
    wb.save(buffer)
    buffer.seek(0)
    return buffer


# ===== ЭНДПОИНТЫ =====

@router.get("/executed")
async def get_executed_report(
    current_user: Annotated[User, Depends(get_current_active_user)],
    session: AsyncSession = Depends(get_session),
) -> StreamingResponse:
    """
    Генерация Excel отчета по всем выполненным заявкам.
    Доступно любому авторизованному пользователю.
    """
    query = (
        select(Ticket)
        .options(selectinload(Ticket.executor))
        .where(Ticket.status == TicketStatus.DONE)
    )
    result = await session.exec(query)
    tickets = result.all()

    buffer = create_executed_xlsx(list(tickets))
    filename = f"executed_tickets_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.get("/{ticket_id}")
async def get_ticket_report(
    ticket_id: int,
    current_user: Annotated[User, Depends(get_current_active_user)],
    format: str = "pdf",
    session: AsyncSession = Depends(get_session),
) -> StreamingResponse:
    """Генерация отчета по конкретной заявке (PDF или XLSX)."""
    ticket = await get_ticket_by_id(session, ticket_id)
    if not ticket:
        raise HTTPException(status_code=404, detail="Заявка не найдена")

    # Проверка прав доступа
    if current_user.role == UserRole.CUSTOMER and ticket.customer_id != current_user.id:
        raise HTTPException(status_code=403, detail="Нет доступа к этой заявке")
    if current_user.role == UserRole.EXECUTOR and ticket.executor_id != current_user.id:
        raise HTTPException(status_code=403, detail="Нет доступа к этой заявке")

    if format.lower() == "pdf":
        buffer = create_ticket_pdf(ticket)
        media_type = "application/pdf"
        filename = f"ticket_{ticket_id}_report.pdf"
    elif format.lower() == "xlsx":
        buffer = create_ticket_xlsx(ticket)
        media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        filename = f"ticket_{ticket_id}_report.xlsx"
    else:
        raise HTTPException(status_code=400, detail="Неподдерживаемый формат")

    return StreamingResponse(
        buffer,
        media_type=media_type,
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.get("/digest/{range}")
async def get_digest_report(
    range: DigestRange,
    current_user: Annotated[User, Depends(check_admin_role)],
    format: str = "pdf",
    session: AsyncSession = Depends(get_session),
) -> StreamingResponse:
    """Генерация сводного отчета за период (только для администраторов)."""
    now = datetime.utcnow()
    if range == DigestRange.DAILY:
        start_date = now.replace(hour=0, minute=0, second=0, microsecond=0)
    else:  # weekly
        start_date = now - timedelta(days=7)

    # Загружаем заявки с необходимыми связями сразу
    query = (
        select(Ticket)
        .options(selectinload(Ticket.customer), selectinload(Ticket.executor))
        .where(Ticket.created_at >= start_date)
    )
    result = await session.exec(query)
    tickets = list(result.unique().all())

    period_name = "daily" if range == DigestRange.DAILY else "weekly"

    if format.lower() == "pdf":
        buffer = create_digest_pdf(tickets, period_name)
        media_type = "application/pdf"
        filename = f"digest_{period_name}_{now.strftime('%Y%m%d')}.pdf"
    elif format.lower() == "xlsx":
        buffer = create_digest_xlsx(tickets, period_name)
        media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        filename = f"digest_{period_name}_{now.strftime('%Y%m%d')}.xlsx"
    else:
        raise HTTPException(status_code=400, detail="Неподдерживаемый формат")

    return StreamingResponse(
        buffer,
        media_type=media_type,
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )